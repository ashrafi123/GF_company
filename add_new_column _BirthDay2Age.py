from openpyxl import load_workbook
from datetime import date

workbook = load_workbook(filename="nazm1.xlsx")
workbook.sheetnames
['Sheet1']

sheet = workbook.active



#print(sheet.title)
#print(sheet["L2"].value)

i=2

for value in sheet.iter_rows(min_row=2, min_col=11,max_col=11, values_only=True):

    #print(value[0])
    birthday = value[0].split('/')
    #print(birthday)

    if (int(birthday[1]) > 12) or (int(birthday[2]) > 31):
        print ('WRONG')
    else:
        today = 22  #date.today()
        age = 1402 - int(birthday[0]) - ((5, 22) < (int(birthday[1]), int(birthday[2])))
        sheet.cell(row=i, column=12).value=age
        workbook.save(filename="nazm1.xlsx")
        i=i+1
